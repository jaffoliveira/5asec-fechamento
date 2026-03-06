import streamlit as st
import pdfplumber
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook
from datetime import date, datetime, timedelta
import re
import io

# ──────────────────────────────────────────────────────────────
# CONFIGURAÇÃO
# ──────────────────────────────────────────────────────────────
STORES = {
    "SIDE":    "West Side / Pompéia",
    "ZONE":    "West Zone / Sonda",
    "PLACE":   "West Place / Girassol",
    "STATION": "West Station / Paulistânia",
}

COLORS = {"SIDE": "#3498db", "ZONE": "#e74c3c", "PLACE": "#2ecc71", "STATION": "#f39c12"}

# Palavras-chave para identificar cada loja no PDF de fechamento
STORE_KEYWORDS = {
    "SIDE":    ["POMPEIA, 1700", "AV. POMPEIA, 1700", "AV POMPEIA 1700", "3672 1466", "3672 1465"],
    "ZONE":    ["CARLOS VICARI", "SONDA POMPEIA", "3675 4295"],
    "PLACE":   ["GIRASSOL", "VILA MADALENA", "3032 3956"],
    "STATION": ["PAULISTANIA", "PAULISTÂNIA", "SUMAREZINHO", "3675 3321"],
}

# Colunas para preencher no template Excel
EXCEL_COL = {
    'pecas': 4, 'servicos': 5, 'fatu': 6, 'apagar': 7,
    'din': 10, 'cheque': 11, 'cc': 12, 'cd': 13,
    'dep': 14, 'ecom': 15, 'outros': 16, 'leitura': 18,
    'agua': 21, 'mercado': 22, 'cafe': 23, 'pedagio': 24,
    'farmacia': 25, 'bolo': 26, 'banco': 27, 'outros_sangria': 28,
    'fundo': 29,
}

# Palavras-chave para classificar lançamentos bancários como sangrias
SANGRIA_KW = {
    'agua':          ['AGUA', 'ÁGUA', 'ACQUA'],
    'mercado':       ['MERCADO', 'SUPERMERCADO', 'HORTIFRUTI', 'ATACADAO', 'ATACADÃO'],
    'cafe':          ['CAFE', 'CAFÉ', 'CAFETERIA'],
    'pedagio':       ['PEDAGIO', 'PEDÁGIO', 'AUTOBAN', 'ECOVIAS', 'VIAOLTRA', 'CONCER', 'ARTERIS'],
    'farmacia':      ['FARMACIA', 'FARMÁCIA', 'DROGARIA', 'DROGASIL', 'DROGA'],
    'bolo':          ['BOLO', 'CONFEITARIA', 'DOCES', 'CAKE'],
    'banco':         ['BUSINESS', 'REDE ', 'GETNET', 'CIELO', 'STONE', 'PAGSEGURO',
                      'PAGBANK', 'MAQUININHA', 'ADQUIRENTE', 'TAXA ', 'TARIFA ',
                      'DEBITO SEGURO', 'SEGURO', 'JUROS', 'IOF'],
}

# ──────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────
def fmt_brl(v):
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return "R$ 0,00"

def safe_float(v, default=0.0):
    try:
        return float(v) if v is not None and str(v).strip() not in ['', '-', 'nan'] else default
    except:
        return default

def get_val(sd, key):
    return safe_float(sd.get(key, 0))

def total_sangria(sd):
    keys = ['agua', 'mercado', 'cafe', 'pedagio', 'farmacia', 'bolo', 'banco', 'outros_sangria']
    return sum(get_val(sd, k) for k in keys)

def total_recebido(sd):
    return sum(get_val(sd, k) for k in ['din', 'cheque', 'cc', 'cd', 'dep', 'ecom', 'outros'])

def empty_store():
    return {k: 0.0 for k in [
        'pecas', 'servicos', 'fatu', 'apagar', 'leitura', 'fundo',
        'din', 'cheque', 'cc', 'cd', 'dep', 'ecom', 'outros',
        'agua', 'mercado', 'cafe', 'pedagio', 'farmacia', 'bolo', 'banco', 'outros_sangria',
        '_cc_ref', '_cd_ref', '_dep_ref',
    ]}

def ni(label, key, val):
    try:
        v = float(val) if val else 0.0
    except:
        v = 0.0
    return st.number_input(label, min_value=0.0, value=v, step=0.01, format="%.2f", key=key)

def map_web_store(loja_name: str):
    """Mapeia nome de loja do Web Gerencial para store_id."""
    n = str(loja_name).upper()
    if 'SONDA' in n:
        return 'ZONE'
    if 'GIRASSOL' in n or 'MADALENA' in n:
        return 'PLACE'
    if 'SUMAREZINHO' in n:
        return 'STATION'
    if 'POMPEIA' in n:
        return 'SIDE'
    return None

def map_rede_store(nome: str):
    """Mapeia nome do estabelecimento para store_id."""
    n = str(nome).upper()
    if 'WEST SIDE' in n:    return 'SIDE'
    if 'WEST ZONE' in n:    return 'ZONE'
    if 'WEST PLACE' in n:   return 'PLACE'
    if 'WEST STATION' in n: return 'STATION'
    if 'SONDA' in n:        return 'ZONE'
    if 'GIRASSOL' in n or 'MADALENA' in n: return 'PLACE'
    if 'SUMAREZINHO' in n:  return 'STATION'
    if 'POMPEIA' in n:      return 'SIDE'
    return None

# ──────────────────────────────────────────────────────────────
# PARSE — PDF fechamento (legado)
# ──────────────────────────────────────────────────────────────
def identify_store_from_pdf(text: str):
    for store_id, keywords in STORE_KEYWORDS.items():
        for kw in keywords:
            if kw.upper() in text.upper():
                return store_id
    return None

def parse_brl(s: str):
    s = str(s).replace(".", "").replace(",", ".").strip()
    try:
        return float(s)
    except:
        return 0.0

def parse_pdf_fechamento(pdf_file):
    results = {}
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            store_id = identify_store_from_pdf(text)
            if not store_id:
                continue
            sd = empty_store()
            patterns = {
                'fatu':    r'(?:FAT\.?\s*L[IÍ]Q\.?|FATURAMENTO\s*L[IÍ]QUIDO)[^\d]*([\d.,]+)',
                'pecas':   r'(?:N[º°]\s*PE[ÇC]AS|PECAS|PE[ÇC]AS)[^\d]*([\d.,]+)',
                'servicos': r'(?:N[º°]\s*SERVI[ÇC]OS|SERVICOS)[^\d]*([\d.,]+)',
                'apagar':  r'(?:A\s*PAGAR|TOTAL\s*A\s*PAGAR)[^\d]*([\d.,]+)',
                'din':     r'(?:DINHEIRO|ESPÉCIE)[^\d]*([\d.,]+)',
                'cc':      r'CRÉDITO[^\d]*([\d.,]+)',
                'cd':      r'DÉBITO[^\d]*([\d.,]+)',
                'dep':     r'(?:DEP[ÓO]SITO|PIX)[^\d]*([\d.,]+)',
                'leitura': r'(?:LEITURA\s*[XZ]|TOTAL\s*VENDAS)[^\d]*([\d.,]+)',
                'fundo':   r'(?:FUNDO\s*(?:DE\s*)?CAIXA|FUNDO\s*ABERTURA)[^\d]*([\d.,]+)',
            }
            for field, pat in patterns.items():
                m = re.search(pat, text, re.IGNORECASE)
                if m:
                    sd[field] = parse_brl(m.group(1))
            results[store_id] = sd
    return results

# ──────────────────────────────────────────────────────────────
# PARSE — Web Gerencial AllegroNet (XLS/XLSX)
# ──────────────────────────────────────────────────────────────
def parse_web_gerencial(file_obj):
    """
    Lê o arquivo Web*.XLS exportado do AllegroNet (Comp. Geral Lojas).
    Retorna dict: { store_id: {pecas, servicos, fatu, tickets, dias, clientes} }
    Suporta .xls (via xlrd) e .xlsx (via openpyxl).
    """
    filename = getattr(file_obj, 'name', '').lower()

    def extract_from_rows(headers, data_rows):
        result = {}
        for row in data_rows:
            if not row or not row[0]:
                continue
            loja = str(row[0]).strip()
            store_id = map_web_store(loja)
            if not store_id:
                continue
            row_vals = {str(headers[c]).strip(): row[c]
                        for c in range(min(len(headers), len(row))) if row[c] is not None}
            result[store_id] = {
                'tickets':  safe_float(row_vals.get('Tickets')),
                'pecas':    safe_float(row_vals.get('Peças',    row_vals.get('Pecas'))),
                'servicos': safe_float(row_vals.get('Serviços', row_vals.get('Servicos'))),
                'fatu':     safe_float(row_vals.get('Faturam.', row_vals.get('Faturamento'))),
                'dias':     safe_float(row_vals.get('Dias')),
                'clientes': safe_float(row_vals.get('Clientes')),
                '_loja':    loja,
            }
        return result

    if filename.endswith('.xls'):
        try:
            import xlrd
            content = file_obj.read()
            wb = xlrd.open_workbook(file_contents=content)
            ws = wb.sheet_by_index(0)
            if ws.nrows < 2:
                return {}
            headers = [ws.cell_value(0, c) for c in range(ws.ncols)]
            data_rows = [[ws.cell_value(r, c) for c in range(ws.ncols)]
                         for r in range(1, ws.nrows)]
            return extract_from_rows(headers, data_rows)
        except ImportError:
            st.warning("⚠️ xlrd não disponível. Converta o arquivo para .xlsx no Excel e tente novamente.")
            return {}
        except Exception as e:
            st.error(f"Erro ao ler XLS: {e}")
            return {}
    else:
        try:
            wb = load_workbook(file_obj, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not rows:
                return {}
            return extract_from_rows(list(rows[0]), list(rows[1:]))
        except Exception as e:
            st.error(f"Erro ao ler gerencial: {e}")
            return {}

# ──────────────────────────────────────────────────────────────
# PARSE — Rede Rel. Vendas (cartões crédito/débito/link)
# ──────────────────────────────────────────────────────────────
def parse_rede_vendas(file_obj):
    """
    Lê Rede_Rel_Vendas*.xlsx.
    Retorna dict: { store_id: {cc, cd, ecom, transactions:[...]} }
    Localiza colunas pelo nome do cabeçalho (robusto a variações de formato).
    """
    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        st.error(f"Erro ao abrir Rede Rel. Vendas: {e}")
        return {}

    # Localiza linha de cabeçalho
    header_row_idx = None
    for i, row in enumerate(rows):
        if row and any(
            str(v).lower().strip() in ('data da venda', 'modalidade', 'valor da venda original')
            for v in row if v
        ):
            header_row_idx = i
            break
    if header_row_idx is None:
        st.warning("Rede Rel. Vendas: cabeçalho não encontrado.")
        return {}

    headers = [str(v).lower().strip() if v else '' for v in rows[header_row_idx]]

    def find_col(keywords):
        for i, h in enumerate(headers):
            if any(kw in h for kw in keywords):
                return i
        return None

    col_data     = find_col(['data da venda'])
    col_status   = find_col(['status da venda', 'status'])
    col_valor    = find_col(['valor da venda original'])
    col_mod      = find_col(['modalidade'])
    col_tipo     = find_col(['tipo'])
    col_parcelas = find_col(['número de parcelas', 'parcelas'])
    col_bandeira = find_col(['bandeira'])
    col_nome     = find_col(['nome do estabelecimento'])

    result = {}
    for row in rows[header_row_idx + 1:]:
        if not row or (col_data is not None and row[col_data] is None):
            continue
        status = str(row[col_status]).lower() if col_status is not None and row[col_status] else ''
        if 'aprovad' not in status:
            continue

        valor    = safe_float(row[col_valor])    if col_valor    is not None and row[col_valor]    is not None else 0.0
        mod      = str(row[col_mod]).lower()     if col_mod      is not None and row[col_mod]      else ''
        tipo     = str(row[col_tipo]).lower()    if col_tipo     is not None and row[col_tipo]     else ''
        bandeira = str(row[col_bandeira])        if col_bandeira is not None and row[col_bandeira] else ''
        nome     = str(row[col_nome])            if col_nome     is not None and row[col_nome]     else ''
        parcelas = int(safe_float(row[col_parcelas], 1)) if col_parcelas is not None and row[col_parcelas] else 1
        data_v   = row[col_data] if col_data is not None else None

        store_id = map_rede_store(nome)
        if not store_id:
            continue

        if store_id not in result:
            result[store_id] = {'cc': 0.0, 'cd': 0.0, 'ecom': 0.0, 'transactions': []}

        if 'débito' in mod or 'debito' in mod:
            cat = 'cd'
        elif 'link' in mod or 'ecommerce' in mod or 'e-commerce' in mod:
            cat = 'ecom'
        else:
            cat = 'cc'

        result[store_id][cat] += valor
        result[store_id]['transactions'].append({
            'data':       data_v,
            'valor':      valor,
            'modalidade': mod,
            'tipo':       tipo,
            'bandeira':   bandeira,
            'parcelas':   parcelas,
            'categoria':  cat,
            'loja':       nome,
            'store_id':   store_id,
        })
    return result

# ──────────────────────────────────────────────────────────────
# PARSE — Extrato bancário
# ──────────────────────────────────────────────────────────────
def classify_extrato_lancamento(lancamento: str, razao: str) -> str:
    text = (str(lancamento) + ' ' + str(razao)).upper()
    for cat, keywords in SANGRIA_KW.items():
        if any(kw in text for kw in keywords):
            return cat
    return 'outros_sangria'

def parse_extrato_lancamentos(file_obj):
    """
    Lê Extrato_Lançamentos*.xlsx.
    Retorna dict: { store_id: {dep, sangrias:{...}, transactions:[...]} }
    """
    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        st.error(f"Erro ao abrir Extrato: {e}")
        return {}

    # Descobre store_id pelo nome da conta
    store_id = None
    filename = getattr(file_obj, 'name', '').upper()
    for row in rows[:12]:
        if row and row[0] and 'nome' in str(row[0]).lower():
            nome_conta = str(row[1]) if len(row) > 1 and row[1] else ''
            store_id = map_rede_store(nome_conta)
            break
    if not store_id:
        for sid in STORES:
            if sid in filename:
                store_id = sid
                break
    if not store_id:
        store_id = 'SIDE'

    # Localiza cabeçalho
    header_row_idx = None
    for i, row in enumerate(rows):
        if row and any(
            str(v).lower().strip() in ('data', 'lançamento', 'lancamento')
            for v in row if v
        ):
            header_row_idx = i
            break
    if header_row_idx is None:
        st.warning("Extrato: cabeçalho não encontrado.")
        return {}

    headers = [str(v).lower().strip() if v else '' for v in rows[header_row_idx]]

    def find_col(keywords):
        for i, h in enumerate(headers):
            if any(kw in h for kw in keywords):
                return i
        return None

    col_data       = find_col(['data'])
    col_lancamento = find_col(['lançamento', 'lancamento'])
    col_razao      = find_col(['razão social', 'razao social', 'razão', 'razao'])
    col_valor      = find_col(['valor'])

    dep_total = 0.0
    sangrias  = {}
    transactions = []

    for row in rows[header_row_idx + 1:]:
        if not row:
            continue
        if col_valor is None or len(row) <= col_valor or row[col_valor] is None:
            continue
        valor = safe_float(row[col_valor])

        lancamento = str(row[col_lancamento]).strip() if col_lancamento is not None and len(row) > col_lancamento and row[col_lancamento] else ''
        razao      = str(row[col_razao]).strip()      if col_razao      is not None and len(row) > col_razao      and row[col_razao]      else ''
        data_str   = row[col_data] if col_data is not None and len(row) > col_data else None

        if 'saldo anterior' in lancamento.lower():
            continue

        if valor > 0:
            cat = 'dep'
            dep_total += valor
        else:
            cat = classify_extrato_lancamento(lancamento, razao)
            sangrias[cat] = sangrias.get(cat, 0) + abs(valor)

        transactions.append({
            'data':         data_str,
            'lancamento':   lancamento,
            'razao_social': razao,
            'valor':        valor,
            'categoria':    cat,
            'store_id':     store_id,
        })

    return {
        store_id: {
            'dep':          dep_total,
            'sangrias':     sangrias,
            'transactions': transactions,
        }
    }

# ──────────────────────────────────────────────────────────────
# PARSE — Excel histórico (template)
# ──────────────────────────────────────────────────────────────
def read_historical_excel(wb, store_id: str):
    if store_id not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[store_id]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        try:
            dt = row[0] if isinstance(row[0], (date, datetime)) else None
            if dt is None:
                continue
            if isinstance(dt, datetime):
                dt = dt.date()
            fatu = safe_float(row[5])
            data.append({'data': dt, 'fatu': fatu, 'store_id': store_id})
        except:
            continue
    return pd.DataFrame(data)

# ──────────────────────────────────────────────────────────────
# SESSION STATE
# ──────────────────────────────────────────────────────────────
if 'data' not in st.session_state:
    st.session_state.data = {s: empty_store() for s in STORES}
if 'hist_wb' not in st.session_state:
    st.session_state.hist_wb = None
if 'rede_transactions' not in st.session_state:
    st.session_state.rede_transactions = []
if 'extrato_transactions' not in st.session_state:
    st.session_state.extrato_transactions = []

# ──────────────────────────────────────────────────────────────
# LAYOUT
# ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="5ASEC Fechamento",
    page_icon="🧺",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
.ref-box {
    background: #f0f4ff; border-left: 3px solid #3498db;
    padding: 6px 10px; border-radius: 4px;
    font-size: 0.82rem; margin-top: 4px;
}
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────
# SIDEBAR
# ──────────────────────────────────────────────────────────────
with st.sidebar:
    st.title("🧺 5ASEC Fechamento")

    st.markdown("#### 📅 Data do fechamento")
    sel_date = st.date_input("", value=date.today(), key="sel_date", format="DD/MM/YYYY")

    st.divider()

    # ── 1. Web Gerencial ──
    st.markdown("#### 📊 Web Gerencial (AllegroNet)")
    st.caption("Comp. Geral Lojas — mesma data nos dois campos")
    web_file = st.file_uploader("Web*.XLS ou .xlsx", type=['xls', 'xlsx'], key='web_upload')
    if web_file:
        if st.button("📥 Importar Gerencial", key='btn_web'):
            with st.spinner("Importando..."):
                web_data = parse_web_gerencial(web_file)
            if web_data:
                imported = []
                for sid, d in web_data.items():
                    sd = st.session_state.data[sid]
                    if d.get('pecas'):    sd['pecas']    = d['pecas']
                    if d.get('servicos'): sd['servicos'] = d['servicos']
                    if d.get('fatu'):     sd['fatu']     = d['fatu']
                    imported.append(STORES[sid].split('/')[0].strip())
                st.success(f"✅ {', '.join(imported)}")
                st.rerun()
            else:
                st.error("Nenhuma loja identificada.")

    st.divider()

    # ── 2. Rede Rel. Vendas ──
    st.markdown("#### 💳 Rede — Rel. Vendas (cartões)")
    st.caption("Crédito, Débito e Link de Pagamento")
    rede_files = st.file_uploader(
        "Rede_Rel_Vendas*.xlsx (1 por loja)",
        type=['xlsx'], accept_multiple_files=True, key='rede_upload'
    )
    if rede_files:
        if st.button("📥 Importar Cartões", key='btn_rede'):
            all_txns = []
            imported = []
            for f in rede_files:
                with st.spinner(f"Lendo {f.name}..."):
                    rede_data = parse_rede_vendas(f)
                for sid, d in rede_data.items():
                    sd = st.session_state.data[sid]
                    sd['cc']   = d.get('cc', 0)
                    sd['cd']   = d.get('cd', 0)
                    sd['ecom'] = d.get('ecom', 0)
                    all_txns.extend(d.get('transactions', []))
                    imported.append(STORES.get(sid, sid))
            st.session_state.rede_transactions = all_txns
            if imported:
                st.success(f"✅ {', '.join(imported)}")
                st.rerun()
            else:
                st.warning("Nenhuma loja identificada.")

    st.divider()

    # ── 3. Extrato bancário ──
    st.markdown("#### 🏦 Extrato Bancário")
    st.caption("Depósitos, PIX recebidos e Sangrias")
    extrato_files = st.file_uploader(
        "Extrato_Lançamentos*.xlsx (1 por loja)",
        type=['xlsx'], accept_multiple_files=True, key='extrato_upload'
    )
    if extrato_files:
        if st.button("📥 Importar Extrato", key='btn_extrato'):
            all_ext = []
            imported = []
            for f in extrato_files:
                with st.spinner(f"Lendo {f.name}..."):
                    ext_data = parse_extrato_lancamentos(f)
                for sid, d in ext_data.items():
                    sd = st.session_state.data[sid]
                    if d.get('dep'):
                        sd['dep'] = d['dep']
                    for cat, amt in d.get('sangrias', {}).items():
                        if cat in sd:
                            sd[cat] = sd.get(cat, 0) + amt
                    all_ext.extend(d.get('transactions', []))
                    imported.append(STORES.get(sid, sid))
            st.session_state.extrato_transactions = all_ext
            if imported:
                st.success(f"✅ {', '.join(imported)}")
                st.rerun()
            else:
                st.warning("Nenhuma loja identificada.")

    st.divider()

    # ── 4. PDFs de fechamento ──
    st.markdown("#### 🖨️ Relatórios de Fechamento (PDF)")
    pdf_files = st.file_uploader(
        "PDFs das lojas (sistema Windows)",
        type=['pdf'], accept_multiple_files=True, key='pdf_upload'
    )
    if pdf_files:
        if st.button("📥 Importar PDFs", key='btn_pdf'):
            for f in pdf_files:
                with st.spinner(f"Lendo {f.name}..."):
                    parsed = parse_pdf_fechamento(f)
                for sid, d in parsed.items():
                    sd = st.session_state.data[sid]
                    for k, v in d.items():
                        if v and v != 0:
                            sd[k] = v
            st.success(f"✅ {len(pdf_files)} PDF(s) importado(s).")
            st.rerun()

    st.divider()

    # ── 5. Template histórico ──
    st.markdown("#### 📁 Histórico (template anual)")
    hist_file = st.file_uploader(
        "Fechamento de Caixa - 2026.xlsx", type=['xlsx'], key='hist_upload'
    )
    if hist_file:
        try:
            st.session_state.hist_wb = load_workbook(hist_file, data_only=True)
            st.success("✅ Histórico carregado.")
        except Exception as e:
            st.error(f"Erro: {e}")

# ──────────────────────────────────────────────────────────────
# TABS
# ──────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs([
    "📝 Entrada de Dados",
    "📊 Dashboard",
    "💾 Exportar",
    "🔄 Conciliação",
])

# ──────────────────────────────────────────────────────────────
# TAB 1 — ENTRADA DE DADOS
# ──────────────────────────────────────────────────────────────
with tab1:
    st.header(f"Fechamento — {sel_date.strftime('%d/%m/%Y')}")

    for store_id in STORES:
        store_label = STORES[store_id]
        sd = st.session_state.data[store_id]

        with st.expander(f"🏪 {store_label}", expanded=True):
            c1, c2, c3, c4 = st.columns(4)

            with c1:
                st.markdown("**🧾 Produção**")
                sd['pecas']    = ni("Nº Peças",    f"{store_id}_pecas",    sd['pecas'])
                sd['servicos'] = ni("Nº Serviços", f"{store_id}_servicos", sd['servicos'])
                st.markdown("**💸 Faturamento**")
                sd['fatu']   = ni("Fat. Líquido",  f"{store_id}_fatu",   sd['fatu'])
                sd['apagar'] = ni("A Pagar (Dia)", f"{store_id}_apagar", sd['apagar'])

            with c2:
                st.markdown("**💳 Recebimentos**")
                sd['din']    = ni("Dinheiro",       f"{store_id}_din",  sd['din'])
                sd['cheque'] = ni("Cheque",          f"{store_id}_chq",  sd['cheque'])
                sd['cc']     = ni("Crédito 🔵",      f"{store_id}_cc",   sd['cc'])
                sd['cd']     = ni("Débito 🔵",       f"{store_id}_cd",   sd['cd'])
                sd['dep']    = ni("Depósito/PIX 🟠", f"{store_id}_dep",  sd['dep'])
                sd['ecom']   = ni("E-commerce",      f"{store_id}_ecom", sd['ecom'])
                sd['outros'] = ni("Outros (Saldo)",  f"{store_id}_out",  sd['outros'])

                refs = [(sd.get('_cc_ref'), "CC"), (sd.get('_cd_ref'), "CD"), (sd.get('_dep_ref'), "Dep")]
                ref_lines = [f"<b>{label}:</b> {fmt_brl(v)}" for v, label in refs if v]
                if ref_lines:
                    st.markdown(
                        '<div class="ref-box">📋 Referência PDF:<br>' + " &nbsp;|&nbsp; ".join(ref_lines) + "</div>",
                        unsafe_allow_html=True
                    )

            with c3:
                st.markdown("**📊 Caixa**")
                sd['leitura'] = ni("Leitura X",     f"{store_id}_lx",    sd['leitura'])
                sd['fundo']   = ni("Fundo Abertura", f"{store_id}_fundo", sd['fundo'])
                st.divider()
                t_san = total_sangria(sd)
                t_rec = total_recebido(sd)
                fatu_v  = get_val(sd, 'fatu')
                fundo_v = get_val(sd, 'fundo')
                din_v   = get_val(sd, 'din')
                saldo_caixa = fundo_v + din_v - t_san
                st.metric("Total Recebido", fmt_brl(t_rec))
                st.metric("Saldo Final Caixa", fmt_brl(saldo_caixa))
                if fatu_v > 0:
                    pct = t_rec / fatu_v * 100
                    st.metric("% Recebido", f"{pct:.1f}%")

            with c4:
                st.markdown("**🪙 Sangrias**")
                sd['agua']           = ni("💧 Água",    f"{store_id}_agua",   sd['agua'])
                sd['mercado']        = ni("🛒 Mercado", f"{store_id}_mkt",    sd['mercado'])
                sd['cafe']           = ni("☕ Café",    f"{store_id}_cafe",   sd['cafe'])
                sd['pedagio']        = ni("🚗 Pedágio", f"{store_id}_ped",    sd['pedagio'])
                sd['farmacia']       = ni("💊 Farmácia",f"{store_id}_farm",   sd['farmacia'])
                sd['bolo']           = ni("🎂 Bolo",    f"{store_id}_bolo",   sd['bolo'])
                sd['banco']          = ni("🏦 Banco",   f"{store_id}_banco",  sd['banco'])
                sd['outros_sangria'] = ni("➕ Outros",  f"{store_id}_outsan", sd['outros_sangria'])
                st.metric("Total Sangrias", fmt_brl(t_san))

# ──────────────────────────────────────────────────────────────
# TAB 2 — DASHBOARD
# ──────────────────────────────────────────────────────────────
with tab2:
    st.header(f"📊 Dashboard Gerencial — {sel_date.strftime('%d/%m/%Y')}")

    tot_fatu   = sum(get_val(st.session_state.data[s], 'fatu')   for s in STORES)
    tot_pecas  = sum(get_val(st.session_state.data[s], 'pecas')  for s in STORES)
    tot_apagar = sum(get_val(st.session_state.data[s], 'apagar') for s in STORES)
    tot_din    = sum(get_val(st.session_state.data[s], 'din')    for s in STORES)
    tot_cc     = sum(get_val(st.session_state.data[s], 'cc')     for s in STORES)
    tot_cd     = sum(get_val(st.session_state.data[s], 'cd')     for s in STORES)
    tot_dep    = sum(get_val(st.session_state.data[s], 'dep')    for s in STORES)
    tot_out    = sum(get_val(st.session_state.data[s], 'outros') for s in STORES)
    tot_rec    = tot_din + tot_cc + tot_cd + tot_dep + tot_out
    ticket_med = tot_fatu / tot_pecas if tot_pecas > 0 else 0
    pct_rec    = tot_rec / tot_fatu * 100 if tot_fatu > 0 else 0

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("🏭 Faturamento Total", fmt_brl(tot_fatu))
    k2.metric("👕 Total Peças", f"{int(tot_pecas):,}".replace(',', '.'))
    k3.metric("💰 Recebido Hoje", fmt_brl(tot_rec),
              delta=f"{pct_rec:.1f}% do faturado" if tot_fatu > 0 else None)
    k4.metric("⏳ A Receber (Dia)", fmt_brl(tot_apagar))
    k5.metric("🎫 Ticket Médio", fmt_brl(ticket_med))

    st.divider()

    chart_rows = []
    for s in STORES:
        sd = st.session_state.data[s]
        chart_rows.append({
            'Loja': STORES[s].split(' / ')[0],
            'Faturamento': get_val(sd, 'fatu'),
            'Recebido': total_recebido(sd),
            'A Pagar': get_val(sd, 'apagar'),
            'Peças': get_val(sd, 'pecas'),
        })
    df_day = pd.DataFrame(chart_rows)

    col_a, col_b = st.columns([3, 2])
    with col_a:
        fig_bar = px.bar(
            df_day.melt(id_vars='Loja', value_vars=['Faturamento', 'Recebido', 'A Pagar']),
            x='Loja', y='value', color='variable', barmode='group',
            title='Faturamento × Recebido × A Pagar por Loja',
            labels={'value': 'R$', 'variable': ''},
            color_discrete_map={'Faturamento': '#3498db', 'Recebido': '#2ecc71', 'A Pagar': '#e74c3c'},
        )
        fig_bar.update_layout(height=380, legend=dict(orientation='h', y=1.12))
        st.plotly_chart(fig_bar, use_container_width=True)

    with col_b:
        pay_labels = ['Dinheiro', 'Crédito 🔵', 'Débito 🔵', 'Dep/PIX 🟠', 'Outros']
        pay_vals   = [tot_din, tot_cc, tot_cd, tot_dep, tot_out]
        filt = [(l, v) for l, v in zip(pay_labels, pay_vals) if v > 0]
        if filt:
            labels_f, vals_f = zip(*filt)
            fig_pie = go.Figure(go.Pie(
                labels=list(labels_f), values=list(vals_f), hole=0.42,
                marker_colors=['#2ecc71', '#3498db', '#9b59b6', '#e67e22', '#95a5a6'],
            ))
            fig_pie.update_layout(title='Formas de Pagamento', height=380,
                                  legend=dict(orientation='h', y=-0.1))
            st.plotly_chart(fig_pie, use_container_width=True)
        else:
            st.info("Preencha os recebimentos para ver o gráfico de pagamentos.")

    st.subheader("📋 Resumo por Loja")
    tbl = []
    for s in STORES:
        sd = st.session_state.data[s]
        tbl.append({
            'Loja': STORES[s],
            'Peças': int(get_val(sd, 'pecas')),
            'Fat. Líquido': fmt_brl(get_val(sd, 'fatu')),
            'A Pagar Dia': fmt_brl(get_val(sd, 'apagar')),
            'Dinheiro': fmt_brl(get_val(sd, 'din')),
            'Crédito 🔵': fmt_brl(get_val(sd, 'cc')),
            'Débito 🔵': fmt_brl(get_val(sd, 'cd')),
            'Dep/PIX 🟠': fmt_brl(get_val(sd, 'dep')),
            'Leitura X': fmt_brl(get_val(sd, 'leitura')),
            'Saldo Caixa': fmt_brl(get_val(sd, 'fundo') + get_val(sd, 'din') - total_sangria(sd)),
        })
    st.dataframe(pd.DataFrame(tbl).set_index('Loja'), use_container_width=True)

    if st.session_state.hist_wb:
        st.divider()
        st.subheader("📈 Tendência — Últimos 30 dias")
        hist_dfs = []
        for s in STORES:
            df_h = read_historical_excel(st.session_state.hist_wb, s)
            if not df_h.empty:
                df_h['loja_label'] = STORES[s].split(' / ')[0]
                hist_dfs.append(df_h)

        if hist_dfs:
            df_all = pd.concat(hist_dfs)
            df_all = df_all[df_all['data'] <= sel_date]
            cutoff = sel_date - timedelta(days=30)
            df_30  = df_all[df_all['data'] >= cutoff]
            if not df_30.empty:
                fig_line = px.line(
                    df_30, x='data', y='fatu', color='loja_label',
                    title='Faturamento diário por loja (últimos 30 dias)',
                    labels={'data': 'Data', 'fatu': 'Faturamento (R$)', 'loja_label': 'Loja'},
                    color_discrete_map={
                        'West Side': '#3498db', 'West Zone': '#e74c3c',
                        'West Place': '#2ecc71', 'West Station': '#f39c12'
                    },
                )
                fig_line.update_layout(height=350)
                st.plotly_chart(fig_line, use_container_width=True)

                df_30 = df_30.copy()
                df_30['semana'] = df_30['data'].apply(lambda d: d.isocalendar()[1])
                df_week = df_30.groupby(['semana', 'loja_label'])['fatu'].sum().reset_index()
                fig_week = px.bar(
                    df_week, x='semana', y='fatu', color='loja_label', barmode='stack',
                    title='Faturamento semanal consolidado',
                    labels={'semana': 'Semana do ano', 'fatu': 'R$', 'loja_label': 'Loja'},
                )
                fig_week.update_layout(height=320)
                st.plotly_chart(fig_week, use_container_width=True)
        else:
            st.info("Sem dados históricos preenchidos no template.")

# ──────────────────────────────────────────────────────────────
# TAB 3 — EXPORTAR
# ──────────────────────────────────────────────────────────────
with tab3:
    st.header("💾 Exportar Fechamento")

    st.subheader("1. Preencher o template Excel")
    template_file = st.file_uploader(
        "Carregue 'Fechamento de Caixa - 2026.xlsx'",
        type=['xlsx'], key='tpl_upload'
    )
    if template_file:
        st.success("✅ Template carregado.")
        if st.button("📥 Gerar Excel Preenchido", type="primary"):
            ref_date = date(2026, 1, 1)
            row_num  = (sel_date - ref_date).days + 2
            try:
                wb = load_workbook(template_file)
                for store_id in STORES:
                    if store_id not in wb.sheetnames:
                        continue
                    ws = wb[store_id]
                    sd = st.session_state.data[store_id]
                    for field, col_idx in EXCEL_COL.items():
                        val = sd.get(field)
                        if val is not None and safe_float(val) != 0:
                            ws.cell(row=row_num, column=col_idx).value = float(val)
                    if store_id == 'SIDE':
                        ws.cell(row=row_num, column=30).value = f'=AC{row_num}+J{row_num}-T{row_num}'
                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                st.download_button(
                    "⬇️ Baixar Excel",
                    data=buf,
                    file_name=f"Fechamento_{sel_date.strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                st.success(f"✅ Linha {row_num} para {sel_date.strftime('%d/%m/%Y')}")
            except Exception as e:
                st.error(f"Erro: {e}")

    st.divider()
    st.subheader("2. Exportar como CSV")
    rows_csv = []
    for s in STORES:
        sd = st.session_state.data[s]
        row_d = {'data': sel_date.isoformat(), 'loja': STORES[s]}
        for f in ['pecas', 'servicos', 'fatu', 'apagar', 'din', 'cheque', 'cc', 'cd',
                  'dep', 'ecom', 'outros', 'leitura', 'fundo',
                  'agua', 'mercado', 'cafe', 'pedagio', 'farmacia', 'bolo', 'banco', 'outros_sangria']:
            row_d[f] = sd.get(f) or ''
        rows_csv.append(row_d)
    csv_str = pd.DataFrame(rows_csv).to_csv(index=False, sep=';', decimal=',')
    st.download_button(
        "⬇️ Baixar CSV",
        data=csv_str.encode('utf-8-sig'),
        file_name=f"Fechamento_{sel_date.strftime('%Y%m%d')}.csv",
        mime='text/csv'
    )

    st.divider()
    st.subheader("3. Limpar sessão")
    if st.button("🗑️ Limpar todos os dados", type="secondary"):
        st.session_state.data = {s: empty_store() for s in STORES}
        st.session_state.rede_transactions = []
        st.session_state.extrato_transactions = []
        st.rerun()

# ──────────────────────────────────────────────────────────────
# TAB 4 — CONCILIAÇÃO
# ──────────────────────────────────────────────────────────────
with tab4:
    st.header(f"🔄 Conciliação — {sel_date.strftime('%d/%m/%Y')}")

    # ── Transações Rede ──
    if st.session_state.rede_transactions:
        st.subheader("💳 Transações de Cartão (Rede)")

        df_rede = pd.DataFrame(st.session_state.rede_transactions)
        cf1, cf2, cf3 = st.columns(3)
        with cf1:
            lojas_opts = ['Todas'] + sorted(df_rede['store_id'].unique().tolist())
            lj_f = st.selectbox("Loja", lojas_opts, key='rede_lj')
        with cf2:
            mods_opts = ['Todas'] + sorted(df_rede['modalidade'].unique().tolist())
            md_f = st.selectbox("Modalidade", mods_opts, key='rede_md')
        with cf3:
            bands_opts = ['Todas'] + sorted(df_rede['bandeira'].unique().tolist())
            bd_f = st.selectbox("Bandeira", bands_opts, key='rede_bd')

        df_r = df_rede.copy()
        if lj_f != 'Todas': df_r = df_r[df_r['store_id']  == lj_f]
        if md_f != 'Todas': df_r = df_r[df_r['modalidade'] == md_f]
        if bd_f != 'Todas': df_r = df_r[df_r['bandeira']   == bd_f]

        k1r, k2r, k3r, k4r = st.columns(4)
        k1r.metric("Transações", len(df_r))
        k2r.metric("Crédito",    fmt_brl(df_r[df_r['categoria']=='cc']['valor'].sum()))
        k3r.metric("Débito",     fmt_brl(df_r[df_r['categoria']=='cd']['valor'].sum()))
        k4r.metric("Link/Ecom",  fmt_brl(df_r[df_r['categoria']=='ecom']['valor'].sum()))

        cols_show = [c for c in ['data', 'loja', 'modalidade', 'tipo', 'bandeira', 'parcelas', 'valor'] if c in df_r.columns]
        df_show = df_r[cols_show].copy()
        if 'valor' in df_show.columns:
            df_show['valor (R$)'] = df_show['valor'].apply(
                lambda v: f"R$ {v:,.2f}".replace(',','X').replace('.',',').replace('X','.')
            )
            df_show = df_show.drop(columns=['valor'])
        st.dataframe(df_show, use_container_width=True, height=320)

        # Gráfico por loja/modalidade
        resumo_rede = df_rede.groupby(['store_id', 'categoria'])['valor'].sum().reset_index()
        resumo_rede['loja'] = resumo_rede['store_id'].map(STORES)
        resumo_rede['modalidade'] = resumo_rede['categoria'].map(
            {'cc': 'Crédito', 'cd': 'Débito', 'ecom': 'Link/E-commerce'}
        )
        if not resumo_rede.empty:
            fig_rede = px.bar(
                resumo_rede, x='loja', y='valor', color='modalidade', barmode='stack',
                title='Vendas por Loja e Modalidade',
                labels={'valor': 'R$', 'modalidade': '', 'loja': 'Loja'},
                color_discrete_map={'Crédito': '#3498db', 'Débito': '#9b59b6', 'Link/E-commerce': '#e67e22'},
            )
            fig_rede.update_layout(height=300)
            st.plotly_chart(fig_rede, use_container_width=True)

        buf_rede = io.BytesIO()
        df_rede.to_excel(buf_rede, index=False)
        buf_rede.seek(0)
        st.download_button(
            "⬇️ Baixar transações Rede (Excel)",
            data=buf_rede,
            file_name=f"Rede_{sel_date.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.info("📂 Importe os arquivos **Rede Rel. Vendas** para ver a conciliação de cartões.")

    st.divider()

    # ── Extrato bancário ──
    if st.session_state.extrato_transactions:
        st.subheader("🏦 Extrato Bancário")

        df_ext = pd.DataFrame(st.session_state.extrato_transactions)
        ce1, ce2 = st.columns(2)
        with ce1:
            lojas_ext = ['Todas'] + sorted(df_ext['store_id'].unique().tolist())
            lj_ext = st.selectbox("Loja", lojas_ext, key='ext_lj')
        with ce2:
            cats_ext = ['Todas'] + sorted(df_ext['categoria'].unique().tolist())
            ct_ext = st.selectbox("Categoria", cats_ext, key='ext_ct')

        df_e = df_ext.copy()
        if lj_ext != 'Todas': df_e = df_e[df_e['store_id']  == lj_ext]
        if ct_ext != 'Todas': df_e = df_e[df_e['categoria'] == ct_ext]

        entradas = df_e[df_e['valor'] > 0]['valor'].sum()
        saidas   = df_e[df_e['valor'] < 0]['valor'].sum()
        ke1, ke2, ke3 = st.columns(3)
        ke1.metric("Entradas (DEP/PIX)", fmt_brl(entradas))
        ke2.metric("Saídas (despesas)", fmt_brl(abs(saidas)))
        ke3.metric("Saldo líquido",     fmt_brl(entradas + saidas))

        cols_ext = [c for c in ['data', 'lancamento', 'razao_social', 'valor', 'categoria'] if c in df_e.columns]
        df_e_show = df_e[cols_ext].copy()
        df_e_show.columns = [c.replace('lancamento', 'Lançamento')
                              .replace('razao_social', 'Razão Social')
                              .replace('categoria', 'Categoria')
                              .replace('data', 'Data')
                              for c in df_e_show.columns]
        if 'valor' in df_e_show.columns:
            df_e_show['Valor (R$)'] = df_e_show['valor'].apply(
                lambda v: f"R$ {v:+,.2f}".replace(',','X').replace('.',',').replace('X','.')
            )
            df_e_show = df_e_show.drop(columns=['valor'])
        st.dataframe(df_e_show, use_container_width=True, height=280)

        df_saidas = df_ext[df_ext['valor'] < 0].copy()
        if not df_saidas.empty:
            resumo_s = df_saidas.groupby('categoria')['valor'].sum().abs().reset_index()
            fig_s = px.pie(resumo_s, values='valor', names='categoria',
                           title='Composição das Saídas Bancárias', hole=0.35)
            fig_s.update_layout(height=320)
            st.plotly_chart(fig_s, use_container_width=True)

        buf_ext = io.BytesIO()
        df_ext.to_excel(buf_ext, index=False)
        buf_ext.seek(0)
        st.download_button(
            "⬇️ Baixar Extrato (Excel)",
            data=buf_ext,
            file_name=f"Extrato_{sel_date.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.info("📂 Importe os arquivos de **Extrato Bancário** para ver os lançamentos.")

    st.divider()

    # ── Confronto Rede × Fechamento ──
    st.subheader("🔗 Confronto: Rede × Fechamento de Caixa")
    conf_rows = []
    for s in STORES:
        sd = st.session_state.data[s]
        rede_txns = [t for t in st.session_state.rede_transactions if t.get('store_id') == s]
        cc_rede = sum(t['valor'] for t in rede_txns if t['categoria'] == 'cc')
        cd_rede = sum(t['valor'] for t in rede_txns if t['categoria'] == 'cd')
        cc_fech = get_val(sd, 'cc')
        cd_fech = get_val(sd, 'cd')
        dif_cc  = cc_fech - cc_rede
        dif_cd  = cd_fech - cd_rede
        conf_rows.append({
            'Loja':            STORES[s].split('/')[0].strip(),
            'Créd. Rede':      fmt_brl(cc_rede),
            'Créd. Fechamento':fmt_brl(cc_fech),
            'Dif. Crédito':    fmt_brl(dif_cc),
            'Déb. Rede':       fmt_brl(cd_rede),
            'Déb. Fechamento': fmt_brl(cd_fech),
            'Dif. Débito':     fmt_brl(dif_cd),
        })
    st.dataframe(pd.DataFrame(conf_rows).set_index('Loja'), use_container_width=True)
    st.caption("Diferença = Fechamento − Rede. Ideal: zero. Positivo = falta lançar na Rede. Negativo = sobra na Rede.")
